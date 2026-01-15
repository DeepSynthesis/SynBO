import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from pathlib import Path
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import column_index_from_string, get_column_letter

from rxnopt.utils.logger import console
from rxnopt.utils.util_func import plot_SMILES, sanitize_filename


class ExcelWriter:
    # Constants for layout and image scaling
    IMG_DISPLAY_SIZE = 100  # Target size for the image in pixels
    PX_TO_COL_WIDTH = 1 / 7.5  # Approx: 1 unit width ≈ 7.5 px
    PX_TO_ROW_HEIGHT = 0.75  # Approx: 1 px ≈ 0.75 points

    def __init__(self, condition_dict, opt_metrics):
        self.condition_dict = condition_dict
        self.opt_metrics = opt_metrics

    def write_to_excel(self, output_df, batch_id, figure_output=[], figure_path=None, save_path=None, filetype="xlsx", transpose=False):
        if filetype == "xlsx":
            wb = Workbook()
            ws = self._create_worksheet(wb, batch_id)

            self._add_data_to_worksheet(ws, output_df, transpose)

            self._auto_adjust_dimensions(ws, output_df, figure_output, transpose)

            if figure_output and figure_path:
                console.log("Exporting with specific figures...", style="green")
                for figure_type in figure_output:
                    if figure_type not in output_df.columns:
                        continue

                    col_idx_in_df = output_df.columns.get_loc(figure_type)

                    if figure_type in self.condition_dict.keys():
                        self._process_figure(ws, figure_type, output_df, figure_path, col_idx_in_df, transpose)
                    else:
                        console.log(
                            f"Figure output '{figure_type}' not in condition types, skipping...",
                            style="yellow",
                        )
            else:
                console.log(
                    "No figure output and path provided, exporting with names...",
                    style="green",
                )

            wb.save(save_path.with_suffix(".xlsx"))
        else:
            raise ValueError("Unknown filetype")

    def _create_worksheet(self, wb, batch_id):
        ws = wb.active
        ws.title = f"optimization in batch {batch_id}"
        return ws

    def _auto_adjust_dimensions(self, ws, output_df, figure_output, transpose):
        MAX_SIZE = 70
        FONT_FACTOR = 1.3

        # Ensure figure_output is a list for checking
        figure_cols = figure_output if figure_output else []

        if not transpose:
            # --- Standard Mode: Adjust Column Widths ---
            for col_idx, col_cells in enumerate(ws.columns, 1):
                column_letter = get_column_letter(col_idx)

                try:
                    col_name = output_df.columns[col_idx - 1]
                except IndexError:
                    col_name = None

                if col_name and col_name in figure_cols:
                    final_width = self.IMG_DISPLAY_SIZE * self.PX_TO_COL_WIDTH
                    ws.column_dimensions[column_letter].width = final_width
                    continue

                max_length = 0
                for cell in col_cells:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass

                adjusted_width = (max_length + 2) * FONT_FACTOR
                final_width = min(adjusted_width, MAX_SIZE)
                ws.column_dimensions[column_letter].width = final_width
        else:  # --- Transposed Mode: Adjust Row Heights and First Column Width ---

            for row_idx, row_cells in enumerate(ws.rows, 1):

                try:
                    row_header = row_cells[0].value
                except:
                    row_header = None

                if row_header and row_header in figure_cols:
                    target_height = self.IMG_DISPLAY_SIZE * self.PX_TO_ROW_HEIGHT
                    ws.row_dimensions[row_idx].height = target_height
                    continue

                max_length = 0
                for cell in row_cells:
                    try:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass

                target_height = 25
                if max_length > 20:
                    target_height = 30

                ws.row_dimensions[row_idx].height = target_height

            max_header_len = 0
            for cell in ws["A"]:
                try:
                    if cell.value:
                        l = len(str(cell.value))
                        if l > max_header_len:
                            max_header_len = l
                except:
                    pass
            ws.column_dimensions["A"].width = (max_header_len + 2) * FONT_FACTOR

            has_images = any(col in figure_cols for col in output_df.columns)
            if has_images:
                img_col_width = self.IMG_DISPLAY_SIZE * self.PX_TO_COL_WIDTH
                for col_idx in range(2, ws.max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    if ws.column_dimensions[col_letter].width < img_col_width:
                        ws.column_dimensions[col_letter].width = img_col_width

    def _add_data_to_worksheet(self, ws, output_df, transpose):
        # Basic Style Configuration

        HEADER_HEIGHT = 35
        ROW_HEIGHT = 25

        header_font = Font(name="Arial", size=16, bold=True)
        data_font = Font(name="Arial", size=14, bold=False)
        alignment = Alignment(horizontal="center", vertical="center")

        original_rows = list(dataframe_to_rows(output_df, index=False, header=True))

        if not transpose:
            # --- Standard Writing ---
            for i, row in enumerate(original_rows):
                row_idx = i + 1
                if i == 0:
                    ws.row_dimensions[row_idx].height = HEADER_HEIGHT
                    current_font = header_font
                else:
                    ws.row_dimensions[row_idx].height = ROW_HEIGHT
                    current_font = data_font

                for j, value in enumerate(row):
                    cell = ws.cell(row=row_idx, column=j + 1, value=value)
                    cell.font = current_font
                    cell.alignment = alignment
        else:
            # --- Transposed Writing ---
            transposed_rows = list(zip(*original_rows))

            for i, row in enumerate(transposed_rows):
                row_idx = i + 1

                ws.row_dimensions[row_idx].height = ROW_HEIGHT

                for j, value in enumerate(row):
                    col_idx = j + 1
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Style Logic:
                    # If it's the first column (j==0), it corresponds to original headers -> Header Style
                    if j == 0:
                        cell.font = header_font
                    else:
                        cell.font = data_font

                    cell.alignment = alignment

    def _process_figure(self, ws, figure_type, output_df, figure_path, col_idx_in_df, transpose):
        """
        :param col_idx_in_df: Column index in the original dataframe (int, 0-based)
        """
        PADDING = 2  # Image padding (pixels)

        if not transpose:
            # ================= Standard Mode (Images in Columns) =================

            excel_col_idx = col_idx_in_df + 1
            column_letter = get_column_letter(excel_col_idx)
            target_col_width = self.IMG_DISPLAY_SIZE * self.PX_TO_COL_WIDTH
            current_width = ws.column_dimensions[column_letter].width

            if current_width < target_col_width:
                ws.column_dimensions[column_letter].width = target_col_width

            cell_w_px = int(ws.column_dimensions[column_letter].width / self.PX_TO_COL_WIDTH)

            target_row_height_pt = self.IMG_DISPLAY_SIZE * self.PX_TO_ROW_HEIGHT
            cell_h_px = int(target_row_height_pt / self.PX_TO_ROW_HEIGHT)

            for i in output_df.index:
                excel_row_idx = i + 2  # Header is row 1 -> data starts at 2
                cell_address = f"{column_letter}{excel_row_idx}"

                ws.row_dimensions[excel_row_idx].height = target_row_height_pt

                img_filename = output_df.loc[i, figure_type]
                mol_SMILES = self._get_mol_SMILES(img_filename, figure_type)
                self._insert_one_image(
                    ws,
                    figure_path,
                    figure_type,
                    img_filename,
                    mol_SMILES,
                    cell_address,
                    excel_col_idx - 1,
                    excel_row_idx - 1,
                    cell_w_px,
                    cell_h_px,
                    PADDING,
                )

        else:
            # ================= Transposed Mode (Images in Rows) =================

            excel_row_idx = col_idx_in_df + 1
            target_row_height_pt = self.IMG_DISPLAY_SIZE * self.PX_TO_ROW_HEIGHT
            ws.row_dimensions[excel_row_idx].height = target_row_height_pt
            cell_h_px = int(target_row_height_pt / self.PX_TO_ROW_HEIGHT)

            target_col_width = self.IMG_DISPLAY_SIZE * self.PX_TO_COL_WIDTH
            cell_w_px = int(target_col_width / self.PX_TO_COL_WIDTH)

            for i in output_df.index:
                excel_col_idx = i + 2
                column_letter = get_column_letter(excel_col_idx)

                current_w = ws.column_dimensions[column_letter].width
                if current_w < target_col_width:
                    ws.column_dimensions[column_letter].width = target_col_width

                cell_address = f"{column_letter}{excel_row_idx}"

                img_filename = output_df.loc[i, figure_type]
                mol_SMILES = self._get_mol_SMILES(img_filename, figure_type)

                self._insert_one_image(
                    ws,
                    figure_path,
                    figure_type,
                    img_filename,
                    mol_SMILES,
                    cell_address,
                    excel_col_idx - 1,
                    excel_row_idx - 1,
                    cell_w_px,
                    cell_h_px,
                    PADDING,
                )

    def _insert_one_image(
        self, ws, figure_path, figure_type, img_filename, mol_SMILES, cell_address, col_idx_0, row_idx_0, cell_w_px, cell_h_px, padding
    ):
        """
        :param img_filename: Actually the value from DataFrame (usually SMILES string)
        """
        if pd.isna(img_filename):
            return

        save_dir = Path(figure_path) / figure_type

        safe_name = sanitize_filename(str(img_filename))
        img_path = save_dir / f"{safe_name}.png"
        image_ready = False

        # 3. Check Logic: Exists -> Generate -> Skip
        if img_path.exists():
            # A. Image exists, use it
            image_ready = True
        else:
            # B. Image does not exist, try plotting from SMILES
            # console.log(f"Generating image for {safe_name}...", style="blue")
            res = plot_SMILES(mol_SMILES, str(save_dir), file_name=img_filename)

            if res.get("success", False):
                # Verify file creation
                if img_path.exists():
                    image_ready = True
            else:
                # C. Plotting failed (Invalid SMILES or not a molecule column)
                image_ready = False

        # 4. If no image is available, return and keep text
        if not image_ready:
            return

        # ================= Image Insertion Logic =================

        # Clear cell text only when inserting an image
        cell = ws[cell_address]
        cell.value = ""
        try:
            img = Image(str(img_path))
            orig_w, orig_h = img.width, img.height

            # --- Calculate Scaling ---
            available_w = cell_w_px - (2 * padding)
            available_h = cell_h_px - (2 * padding)

            if orig_w == 0 or orig_h == 0:
                return

            scale_w = available_w / orig_w
            scale_h = available_h / orig_h
            scale = min(scale_w, scale_h)

            new_w = int(orig_w * scale)
            new_h = int(orig_h * scale)

            img.width = new_w
            img.height = new_h

            # --- Centering ---
            offset_x_px = max(0, (cell_w_px - new_w) // 2)
            offset_y_px = max(0, (cell_h_px - new_h) // 2)

            # --- Set Anchor ---
            marker = AnchorMarker(col=col_idx_0, colOff=pixels_to_EMU(offset_x_px), row=row_idx_0, rowOff=pixels_to_EMU(offset_y_px))
            size = XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(new_h))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception as e:
            # If insertion fails (e.g., corrupt image), restore text value
            cell.value = img_filename

    def _get_mol_SMILES(self, img_filename, figure_type):
        if isinstance(self.condition_dict[figure_type], dict):
            try:
                mol_SMILES = self.condition_dict[figure_type][img_filename]
            except:
                mol_SMILES = img_filename
        else:
            mol_SMILES = img_filename

        return mol_SMILES
